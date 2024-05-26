import openpyxl
import statistics
from func_for_hw6 import get_iter_rows,mean_of_row

wb = openpyxl.load_workbook(filename = 'wprcl2.xlsx')
ws = wb.active
ws['T1'] = 'w_mean'



def main():
    for row in ws.iter_rows(min_row=2):
          mean_of_row(ws,row,wb) # find the mean for the each row and add it to the file

main()
        

    