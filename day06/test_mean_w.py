import pytest
from func_for_hw6 import mean_of_row
import openpyxl
from func_for_hw6 import get_iter_rows,mean_of_row_for_test

def test_mean():
    wb = openpyxl.load_workbook(filename = 'for_test.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        result = mean_of_row_for_test(ws,row)
        assert(round(result,6) == 1.135039)